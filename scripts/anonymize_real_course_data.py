from pathlib import Path
from datetime import datetime
from io import BytesIO
import hashlib
import json
import re
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

RAW_DIR = ROOT / "data" / "raw" / "real_courses"
ANON_DIR = ROOT / "data" / "processed" / "real_courses" / "anonymized"
INVENTORY_DIR = ROOT / "data" / "processed" / "real_courses" / "inventory"

ANON_DIR.mkdir(parents=True, exist_ok=True)
INVENTORY_DIR.mkdir(parents=True, exist_ok=True)

REPORT_FILE = INVENTORY_DIR / "anonymization_report.md"
MAPPING_FILE = ANON_DIR / "student_mapping_safe.json"


STUDENT_ID_KEYWORDS = [
    "学号", "学生号", "student_id", "student id", "student number", "student no", "number"
]

NAME_KEYWORDS = [
    "姓名", "名字", "学生姓名", "name", "student name"
]

ACCOUNT_KEYWORDS = [
    "账号", "用户", "用户名", "user", "username", "account"
]

CONTACT_KEYWORDS = [
    "手机号", "手机", "电话", "邮箱", "email", "phone", "mobile"
]

IDENTITY_KEYWORDS = [
    "身份证", "证件", "identity", "id card", "passport"
]


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_column_name(value):
    return normalize_text(value).lower()


def contains_keyword(column_name, keywords):
    text = normalize_column_name(column_name)
    return any(keyword.lower() in text for keyword in keywords)


def column_category(column_name):
    if contains_keyword(column_name, STUDENT_ID_KEYWORDS):
        return "student_id"

    if contains_keyword(column_name, NAME_KEYWORDS):
        return "name"

    if contains_keyword(column_name, ACCOUNT_KEYWORDS):
        return "account"

    if contains_keyword(column_name, CONTACT_KEYWORDS):
        return "contact"

    if contains_keyword(column_name, IDENTITY_KEYWORDS):
        return "identity"

    return "normal"


def is_sensitive_column(column_name):
    return column_category(column_name) != "normal"


def safe_filename(name):
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", name)


class StudentAnonymizer:
    def __init__(self):
        self.mapping = {}

    def get_anonymous_id(self, raw_key):
        raw_key = normalize_text(raw_key)

        if not raw_key:
            raw_key = "UNKNOWN"

        if raw_key not in self.mapping:
            self.mapping[raw_key] = f"S{len(self.mapping) + 1:04d}"

        return self.mapping[raw_key]

    def export_safe_mapping(self):
        safe_mapping = {}

        for raw_key, anonymous_id in self.mapping.items():
            raw_hash = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
            safe_mapping[anonymous_id] = {
                "raw_key_sha256": raw_hash
            }

        MAPPING_FILE.write_text(
            json.dumps(safe_mapping, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )


def get_headers(sheet):
    if sheet.max_row < 1:
        return []

    return [
        normalize_text(cell.value)
        for cell in sheet[1]
    ]


def get_sensitive_column_indexes(headers):
    sensitive = []

    for index, column in enumerate(headers, start=1):
        category = column_category(column)

        if category != "normal":
            sensitive.append({
                "index": index,
                "name": column,
                "category": category
            })

    return sensitive


def get_row_student_key(sheet, row_index, sensitive_columns):
    priority = ["student_id", "account", "name"]

    for category in priority:
        for column in sensitive_columns:
            if column["category"] == category:
                value = sheet.cell(row=row_index, column=column["index"]).value
                value = normalize_text(value)

                if value:
                    return value

    return f"{sheet.title}_row_{row_index}"


def anonymize_sheet(sheet, anonymizer):
    headers = get_headers(sheet)
    sensitive_columns = get_sensitive_column_indexes(headers)

    stats = {
        "sheet_name": sheet.title,
        "rows": max(sheet.max_row - 1, 0),
        "columns": len(headers),
        "sensitive_columns": sensitive_columns,
        "rows_processed": 0
    }

    if not sensitive_columns:
        return stats

    for row_index in range(2, sheet.max_row + 1):
        student_key = get_row_student_key(sheet, row_index, sensitive_columns)
        anonymous_id = anonymizer.get_anonymous_id(student_key)

        for column in sensitive_columns:
            cell = sheet.cell(row=row_index, column=column["index"])

            if column["category"] in ["student_id", "name", "account"]:
                cell.value = anonymous_id

            elif column["category"] in ["contact", "identity"]:
                cell.value = "[removed]"

            else:
                cell.value = anonymous_id

        stats["rows_processed"] += 1

    return stats


def anonymize_workbook_from_path(path, output_path, anonymizer):
    workbook = load_workbook(path)
    return anonymize_loaded_workbook(workbook, output_path, anonymizer)


def anonymize_workbook_from_bytes(content, output_path, anonymizer):
    workbook = load_workbook(BytesIO(content))
    return anonymize_loaded_workbook(workbook, output_path, anonymizer)


def anonymize_loaded_workbook(workbook, output_path, anonymizer):
    workbook_stats = []

    for sheet in workbook.worksheets:
        stats = anonymize_sheet(sheet, anonymizer)
        workbook_stats.append(stats)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return workbook_stats


def anonymize_xlsx_file(path, anonymizer):
    output_path = ANON_DIR / f"{path.stem}_anonymized.xlsx"

    sheet_stats = anonymize_workbook_from_path(
        path,
        output_path,
        anonymizer
    )

    return {
        "source": str(path.name),
        "output": str(output_path.relative_to(ROOT)),
        "type": "xlsx",
        "sheets": sheet_stats
    }


def anonymize_zip_file(path, anonymizer):
    zip_output_dir = ANON_DIR / path.stem
    zip_output_dir.mkdir(parents=True, exist_ok=True)

    results = []

    with zipfile.ZipFile(path, "r") as archive:
        for entry in archive.namelist():
            if entry.endswith("/"):
                continue

            if not entry.lower().endswith(".xlsx"):
                continue

            content = archive.read(entry)
            output_name = safe_filename(Path(entry).stem) + "_anonymized.xlsx"
            output_path = zip_output_dir / output_name

            try:
                sheet_stats = anonymize_workbook_from_bytes(
                    content,
                    output_path,
                    anonymizer
                )

                results.append({
                    "source": f"{path.name}::{entry}",
                    "output": str(output_path.relative_to(ROOT)),
                    "type": "xlsx_inside_zip",
                    "sheets": sheet_stats
                })

            except Exception as error:
                results.append({
                    "source": f"{path.name}::{entry}",
                    "type": "xlsx_inside_zip",
                    "error": str(error)
                })

    return {
        "source": str(path.name),
        "type": "zip",
        "entries_processed": results
    }


def write_report(results, anonymizer):
    lines = []

    lines.append("# Real Course Data Anonymization Report")
    lines.append("")
    lines.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## Purpose")
    lines.append("")
    lines.append(
        "This report describes the anonymization process applied to the real course files. "
        "It does not contain raw student names or raw student numbers."
    )
    lines.append("")
    lines.append("## Rules applied")
    lines.append("")
    lines.append("- Student identifiers, names and account fields are replaced with anonymous IDs such as `S0001`.")
    lines.append("- Phone numbers, emails and identity fields are replaced with `[removed]`.")
    lines.append("- Raw files remain local in `data/raw/real_courses/` and must never be committed.")
    lines.append("- Anonymized Excel files remain local in `data/processed/real_courses/anonymized/`.")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Anonymous students detected: **{len(anonymizer.mapping)}**")
    lines.append(f"- Safe mapping file: `{MAPPING_FILE.relative_to(ROOT)}`")
    lines.append("")

    lines.append("## Files processed")
    lines.append("")

    for result in results:
        lines.append(f"### {result['source']}")
        lines.append("")
        lines.append(f"- Type: `{result['type']}`")

        if "output" in result:
            lines.append(f"- Output: `{result['output']}`")

        if "error" in result:
            lines.append(f"- Error: `{result['error']}`")
            lines.append("")
            continue

        if result["type"] == "zip":
            entries = result.get("entries_processed", [])
            lines.append(f"- Excel entries processed: **{len(entries)}**")
            lines.append("")

            for entry in entries:
                lines.append(f"#### {entry['source']}")
                if "output" in entry:
                    lines.append(f"- Output: `{entry['output']}`")
                if "error" in entry:
                    lines.append(f"- Error: `{entry['error']}`")
                if "sheets" in entry:
                    write_sheet_stats(lines, entry["sheets"])
                lines.append("")

        elif "sheets" in result:
            write_sheet_stats(lines, result["sheets"])

        lines.append("")

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")


def write_sheet_stats(lines, sheets):
    for sheet in sheets:
        lines.append(f"#### Sheet: `{sheet['sheet_name']}`")
        lines.append(f"- Rows: **{sheet['rows']}**")
        lines.append(f"- Columns: **{sheet['columns']}**")
        lines.append(f"- Rows processed: **{sheet['rows_processed']}**")

        sensitive_columns = sheet["sensitive_columns"]

        if sensitive_columns:
            lines.append("- Sensitive columns anonymized:")
            for column in sensitive_columns:
                lines.append(f"  - `{column['name']}` → {column['category']}")
        else:
            lines.append("- Sensitive columns anonymized: none detected")

        lines.append("")


def main():
    if not RAW_DIR.exists():
        raise FileNotFoundError(f"Raw directory not found: {RAW_DIR}")

    files = [
        path for path in RAW_DIR.iterdir()
        if path.is_file()
    ]

    anonymizer = StudentAnonymizer()
    results = []

    for path in files:
        print(f"Processing {path.name}...")

        try:
            if path.suffix.lower() == ".xlsx":
                results.append(anonymize_xlsx_file(path, anonymizer))

            elif path.suffix.lower() == ".zip":
                results.append(anonymize_zip_file(path, anonymizer))

            else:
                results.append({
                    "source": path.name,
                    "type": path.suffix.lower().replace(".", "") or "unknown",
                    "note": "Skipped: unsupported file type for anonymization."
                })

        except Exception as error:
            results.append({
                "source": path.name,
                "type": path.suffix.lower().replace(".", ""),
                "error": str(error)
            })

    anonymizer.export_safe_mapping()
    write_report(results, anonymizer)

    print("")
    print(f"Anonymization report written to: {REPORT_FILE}")
    print(f"Anonymized files written to: {ANON_DIR}")
    print(f"Anonymous students detected: {len(anonymizer.mapping)}")


if __name__ == "__main__":
    main()