"""
Génère data/class_export.xlsx avec toutes les colonnes requises par l'ETL.

Feuilles produites :
  - 学生学习进度详情   (student progress)
  - 综合成绩          (composite grades)

Colonnes ETL attendues (index 0-based) :
  学生学习进度详情
    0  external_uid
    1  full_name
    2  student_no      ← clé
    3  school
    4  department
    5  major
    6  class_name
    7  admission_year
    8  tasks           "done/total"
    9  (vide)
   10  (vide)
   11  discussion_count
   12  chapter_visit_count
   13  learning_status

  综合成绩
    0  seq
    1  full_name
    2  student_no      ← clé
    3  school
    4  department
    5  major
    6  class_name
    7  homework_score  (0-100)
    8  final_score     (0-100)

Header rows : 0,1 = titre fusionné, 2 = en-têtes, data à partir de 3.
"""
import random
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installation de openpyxl…")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

OUT = Path(__file__).resolve().parents[1] / "data" / "class_export.xlsx"

STUDENTS = [
    ("U001", "张伟",   "20240001", "工学院", "计算机系", "人工智能", "AI2401", 2024),
    ("U002", "李娜",   "20240002", "工学院", "计算机系", "人工智能", "AI2401", 2024),
    ("U003", "王芳",   "20240003", "工学院", "软件工程", "软件工程", "SE2401", 2024),
    ("U004", "刘洋",   "20240004", "工学院", "软件工程", "软件工程", "SE2401", 2024),
    ("U005", "陈静",   "20240005", "工学院", "计算机系", "数据科学", "DS2401", 2024),
    ("U006", "杨磊",   "20240006", "工学院", "计算机系", "数据科学", "DS2401", 2024),
    ("U007", "赵敏",   "20240007", "工学院", "软件工程", "人工智能", "AI2401", 2024),
    ("U008", "黄佳",   "20240008", "工学院", "计算机系", "人工智能", "AI2401", 2024),
    ("U009", "周鑫",   "20240009", "工学院", "软件工程", "软件工程", "SE2401", 2024),
    ("U010", "吴雪",   "20240010", "工学院", "计算机系", "数据科学", "DS2401", 2024),
]


def r_score(mu, sigma=8):
    return round(max(40, min(100, random.gauss(mu, sigma))), 1)


wb = openpyxl.Workbook()

# ── 学生学习进度详情 ─────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "学生学习进度详情"

# Rows 0-1 : titre (row index 1-2 in openpyxl 1-based)
ws1.merge_cells("A1:N1")
ws1["A1"] = "人工智能基础 - 学生学习进度详情"
ws1["A1"].font = Font(bold=True, size=13)
ws1["A1"].alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:N2")
ws1["A2"] = "课程编号：AI101   学期：2024-春"
ws1["A2"].alignment = Alignment(horizontal="center")

# Row 3 : headers (openpyxl row 3)
headers_sp = [
    "外部UID", "姓名", "学号", "学院", "院系", "专业", "班级", "入学年份",
    "任务完成/总数", "", "", "讨论次数", "章节访问数", "学习状态"
]
for col, h in enumerate(headers_sp, start=1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

# Data rows from row 4
for s in STUDENTS:
    uid, name, sno, school, dept, major, cls, year = s
    done = random.randint(3, 10)
    total = 10
    disc = random.randint(0, 15)
    visits = random.randint(5, 40)
    status = random.choice(["学习中", "学习中", "活跃", "不活跃"])
    row = [uid, name, sno, school, dept, major, cls, year,
           f"{done}/{total}", "", "", disc, visits, status]
    ws1.append(row)

# ── 综合成绩 ─────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("综合成绩")

ws2.merge_cells("A1:I1")
ws2["A1"] = "人工智能基础 - 综合成绩"
ws2["A1"].font = Font(bold=True, size=13)
ws2["A1"].alignment = Alignment(horizontal="center")

ws2.merge_cells("A2:I2")
ws2["A2"] = "课程编号：AI101   学期：2024-春"
ws2["A2"].alignment = Alignment(horizontal="center")

headers_cg = ["序号", "姓名", "学号", "学院", "院系", "专业", "班级", "作业成绩(100%)", "综合成绩"]
for col, h in enumerate(headers_cg, start=1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="E2EFDA")

for i, s in enumerate(STUDENTS, start=1):
    uid, name, sno, school, dept, major, cls, year = s
    hw = r_score(75)
    final = r_score(72)
    ws2.append([i, name, sno, school, dept, major, cls, hw, final])

# ── Sauvegarde ───────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"✅ Fichier généré : {OUT}")
print(f"   {len(STUDENTS)} étudiants")
print(f"   Feuilles : {wb.sheetnames}")
print()
print("Comptes qui seront créés après upload (login = mot de passe initial) :")
for s in STUDENTS:
    print(f"  {s[2]}  ({s[1]})")
