from pathlib import Path
from datetime import datetime
import csv
import hashlib
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

ANON_DIR = ROOT / "data" / "processed" / "real_courses" / "anonymized"
NORMALIZED_DIR = ROOT / "data" / "processed" / "real_courses" / "normalized"
INVENTORY_DIR = ROOT / "data" / "processed" / "real_courses" / "inventory"

NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)
INVENTORY_DIR.mkdir(parents=True, exist_ok=True)

REPORT_FILE = INVENTORY_DIR / "normalization_report.md"


OUTPUT_FILES = {
    "students": NORMALIZED_DIR / "students.csv",
    "subjects": NORMALIZED_DIR / "subjects.csv",
    "works": NORMALIZED_DIR / "works.csv",
    "answers": NORMALIZED_DIR / "answers.csv",
    "activities": NORMALIZED_DIR / "activities.csv",
    "concepts": NORMALIZED_DIR / "concepts.csv",
    "relations": NORMALIZED_DIR / "relations.csv",
    "resources": NORMALIZED_DIR / "resources.csv",
}


CONCEPT_CATALOG = [
    {
        "concept_id": "C_AI_OVERVIEW",
        "concept_name": "AI overview",
        "chapter": "AI fundamentals",
        "description": "General understanding of what artificial intelligence is.",
        "keywords": ["人工智能概论", "什么是人工智能", "人工智能", "ai introduction", "ai overview"]
    },
    {
        "concept_id": "C_AI_HISTORY",
        "concept_name": "AI history",
        "chapter": "AI fundamentals",
        "description": "Origins and historical development of artificial intelligence.",
        "keywords": ["起源", "发展", "历史", "history", "origin"]
    },
    {
        "concept_id": "C_SYMBOLIC_AI",
        "concept_name": "Symbolic AI",
        "chapter": "AI paradigms",
        "description": "Symbolic reasoning and rule-based approaches.",
        "keywords": ["符号主义", "symbolic"]
    },
    {
        "concept_id": "C_KNOWLEDGE_REPRESENTATION",
        "concept_name": "Knowledge representation",
        "chapter": "Knowledge engineering",
        "description": "How knowledge can be formally represented.",
        "keywords": ["知识表示", "knowledge representation"]
    },
    {
        "concept_id": "C_KNOWLEDGE_GRAPH",
        "concept_name": "Knowledge graph",
        "chapter": "Knowledge engineering",
        "description": "Structured representation of entities and relations.",
        "keywords": ["知识图谱", "knowledge graph"]
    },
    {
        "concept_id": "C_MACHINE_LEARNING_AWARENESS",
        "concept_name": "Machine learning awareness",
        "chapter": "AI paradigms",
        "description": "General awareness of machine learning ideas.",
        "keywords": ["机器学习", "machine learning", "ml"]
    },
    {
        "concept_id": "C_NEURAL_NETWORK_AWARENESS",
        "concept_name": "Neural network awareness",
        "chapter": "AI paradigms",
        "description": "General awareness of neural networks.",
        "keywords": ["神经网络", "neural network"]
    },
    {
        "concept_id": "C_DEEP_LEARNING_AWARENESS",
        "concept_name": "Deep learning awareness",
        "chapter": "AI paradigms",
        "description": "General awareness of deep learning.",
        "keywords": ["深度学习", "deep learning", "dl"]
    },
    {
        "concept_id": "C_NLP",
        "concept_name": "Natural language processing",
        "chapter": "AI applications",
        "description": "AI techniques applied to language.",
        "keywords": ["自然语言", "语言处理", "nlp"]
    },
    {
        "concept_id": "C_LLM",
        "concept_name": "Large language models",
        "chapter": "Generative AI",
        "description": "General understanding of large language models.",
        "keywords": ["大模型", "语言模型", "llm", "chatgpt", "deepseek"]
    },
    {
        "concept_id": "C_GENERATIVE_AI",
        "concept_name": "Generative AI",
        "chapter": "Generative AI",
        "description": "AI systems able to generate text, image or other content.",
        "keywords": ["生成式", "生成", "aigc", "generative"]
    },
    {
        "concept_id": "C_COMPUTER_VISION",
        "concept_name": "Computer vision",
        "chapter": "AI applications",
        "description": "AI techniques applied to images and visual data.",
        "keywords": ["图像", "视觉", "computer vision", "cv"]
    },
    {
        "concept_id": "C_AI_APPLICATIONS",
        "concept_name": "AI applications",
        "chapter": "AI applications",
        "description": "Practical uses of AI in society and industry.",
        "keywords": ["应用", "案例", "application", "case"]
    },
    {
        "concept_id": "C_AI_ETHICS",
        "concept_name": "AI ethics and risks",
        "chapter": "AI society",
        "description": "Ethical, social, privacy and risk issues related to AI.",
        "keywords": ["伦理", "风险", "安全", "隐私", "ethics", "risk", "privacy", "safety"]
    },
]


RELATION_CATALOG = [
    ("REL001", "C_SYMBOLIC_AI", "C_AI_OVERVIEW", "requires", "Symbolic AI requires a general understanding of AI."),
    ("REL002", "C_KNOWLEDGE_REPRESENTATION", "C_SYMBOLIC_AI", "requires", "Knowledge representation is part of symbolic AI."),
    ("REL003", "C_KNOWLEDGE_GRAPH", "C_KNOWLEDGE_REPRESENTATION", "requires", "Knowledge graph requires knowledge representation."),
    ("REL004", "C_MACHINE_LEARNING_AWARENESS", "C_AI_OVERVIEW", "requires", "Machine learning requires a general understanding of AI."),
    ("REL005", "C_NEURAL_NETWORK_AWARENESS", "C_MACHINE_LEARNING_AWARENESS", "requires", "Neural networks are introduced after machine learning."),
    ("REL006", "C_DEEP_LEARNING_AWARENESS", "C_NEURAL_NETWORK_AWARENESS", "requires", "Deep learning requires basic neural network awareness."),
    ("REL007", "C_LLM", "C_NLP", "requires", "Large language models are linked to language processing."),
    ("REL008", "C_GENERATIVE_AI", "C_LLM", "requires", "Generative AI examples often rely on large language models."),
    ("REL009", "C_AI_ETHICS", "C_AI_APPLICATIONS", "related_to", "Ethical risks are linked to AI applications."),
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value):
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        return round(float(value), 4)

    text = clean(value)
    if not text:
        return ""

    text = text.replace("%", "").replace(",", ".")
    match = re.search(r"-?\d+(\.\d+)?", text)

    if not match:
        return ""

    try:
        return round(float(match.group(0)), 4)
    except ValueError:
        return ""


def parse_rate(value):
    number = parse_number(value)

    if number == "":
        return ""

    if number > 1:
        return round(number / 100, 4)

    return round(number, 4)


def stable_id(prefix, text, length=10):
    raw = clean(text).encode("utf-8")
    digest = hashlib.sha1(raw).hexdigest()[:length].upper()
    return f"{prefix}_{digest}"


def is_anonymous_student_id(value):
    text = clean(value)
    return bool(re.fullmatch(r"S\d{4,}", text))


def get_cell(sheet, row, col):
    return clean(sheet.cell(row=row, column=col).value)


def get_metadata(sheet):
    text = get_cell(sheet, 2, 1)

    course_name = "Unknown subject"
    class_name = ""
    teacher = ""

    course_match = re.search(r"课程[:：]\s*(.*?)\s{2,}", text)
    class_match = re.search(r"班级[:：]\s*(.*?)\s{2,}", text)
    teacher_match = re.search(r"任课教师[:：]\s*(.*?)\s{1,}", text)

    if course_match:
        course_name = course_match.group(1).strip()

    if class_match:
        class_name = class_match.group(1).strip()

    if teacher_match:
        teacher = teacher_match.group(1).strip()

    return {
        "course_name": course_name,
        "class_name": class_name,
        "teacher": teacher
    }


def find_column(sheet, header_row, candidates):
    for col in range(1, sheet.max_column + 1):
        value = get_cell(sheet, header_row, col)

        for candidate in candidates:
            if candidate in value:
                return col

    return None


def detect_student_id(sheet, row, header_row=3):
    id_col = find_column(sheet, header_row, ["学号", "学生号", "UID", "student"])
    name_col = find_column(sheet, header_row, ["学生姓名", "姓名", "name"])

    if id_col:
        value = get_cell(sheet, row, id_col)
        if is_anonymous_student_id(value):
            return value

    if name_col:
        value = get_cell(sheet, row, name_col)
        if is_anonymous_student_id(value):
            return value

    return ""


def detect_group(sheet, row, header_row=3):
    group_col = find_column(sheet, header_row, ["班级", "行政班级", "class"])
    if group_col:
        return get_cell(sheet, row, group_col)
    return ""


def detect_major(sheet, row, header_row=3):
    major_col = find_column(sheet, header_row, ["专业", "major"])
    if major_col:
        return get_cell(sheet, row, major_col)
    return ""


def normalize_title(text):
    text = clean(text)
    text = re.sub(r"\s+", " ", text)
    return text


def concept_for_text(text):
    normalized = clean(text).lower()

    for concept in CONCEPT_CATALOG:
        for keyword in concept["keywords"]:
            if keyword.lower() in normalized:
                return concept["concept_id"]

    return "C_AI_OVERVIEW"


def resource_type_for_title(title):
    title = clean(title)

    if ".mp4" in title.lower() or "视频" in title:
        return "video"

    if ".pdf" in title.lower() or "资料" in title:
        return "document"

    if "测试" in title or "quiz" in title.lower():
        return "quiz"

    if "作业" in title or "homework" in title.lower():
        return "homework"

    return "learning_task"


def work_type_for_sheet(sheet_name):
    if sheet_name == "作业统计":
        return "homework"

    if sheet_name == "考试统计":
        return "exam"

    if sheet_name == "章节测验统计":
        return "quiz"

    if sheet_name == "线下成绩统计":
        return "offline"

    if "综合成绩" in sheet_name:
        return "overall"

    return "work"


def is_completed_from_status(status):
    status = clean(status)

    if not status:
        return ""

    if status in ["已完成", "已提交", "已签", "教师代签", "完成"]:
        return "true"

    if status in ["未完成", "未提交", "缺勤", "未参与"]:
        return "false"

    return ""


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()

        for row in rows:
            clean_row = {field: row.get(field, "") for field in fieldnames}
            writer.writerow(clean_row)


class NormalizedData:
    def __init__(self):
        self.students = {}
        self.subjects = {}
        self.works = {}
        self.answers = {}
        self.activities = {}
        self.resources = {}
        self.concepts = {}
        self.relations = {}

    def add_subject(self, source_file, metadata):
        subject_key = f"{source_file}::{metadata['course_name']}::{metadata['class_name']}"
        subject_id = stable_id("SUBJ", subject_key)

        if subject_id not in self.subjects:
            self.subjects[subject_id] = {
                "subject_id": subject_id,
                "subject_name": metadata["course_name"],
                "teacher": metadata["teacher"],
                "class_name": metadata["class_name"],
                "source_file": source_file,
                "description": "Real course export normalized from anonymized teaching platform data."
            }

        return subject_id

    def add_student(self, student_id, subject_id, group_name="", major="", source_file=""):
        if not student_id:
            return

        key = f"{student_id}::{subject_id}"

        if key not in self.students:
            self.students[key] = {
                "student_id": student_id,
                "subject_id": subject_id,
                "group_name": group_name,
                "major": major,
                "source_file": source_file
            }

    def add_work(self, subject_id, title, work_type, source_sheet, max_score=100, date=""):
        title = normalize_title(title)

        if not title:
            return ""

        work_id = stable_id("W", f"{subject_id}::{source_sheet}::{title}::{work_type}")

        if work_id not in self.works:
            self.works[work_id] = {
                "work_id": work_id,
                "subject_id": subject_id,
                "work_title": title,
                "work_type": work_type,
                "date": date,
                "max_score": max_score,
                "source_sheet": source_sheet
            }

        return work_id

    def add_answer(
        self,
        student_id,
        subject_id,
        work_id,
        concept_id,
        score="",
        max_score=100,
        status="",
        submitted_at="",
        source_sheet=""
    ):
        if not student_id or not work_id:
            return

        score_value = parse_number(score)
        score_rate = ""

        if score_value != "" and max_score:
            try:
                score_rate = round(float(score_value) / float(max_score), 4)
            except ZeroDivisionError:
                score_rate = ""

        is_completed = is_completed_from_status(status)

        is_correct = ""
        if score_rate != "":
            is_correct = "true" if score_rate >= 0.6 else "false"

        answer_id = stable_id(
            "A",
            f"{student_id}::{subject_id}::{work_id}::{concept_id}::{source_sheet}"
        )

        self.answers[answer_id] = {
            "answer_id": answer_id,
            "student_id": student_id,
            "subject_id": subject_id,
            "work_id": work_id,
            "concept_id": concept_id,
            "score": score_value,
            "max_score": max_score,
            "score_rate": score_rate,
            "is_correct": is_correct,
            "status": status,
            "is_completed": is_completed,
            "submitted_at": submitted_at,
            "source_sheet": source_sheet
        }

    def add_resource(self, subject_id, title, source_sheet):
        title = normalize_title(title)

        if not title:
            return ""

        concept_id = concept_for_text(title)
        resource_type = resource_type_for_title(title)

        resource_id = stable_id("R", f"{subject_id}::{title}::{resource_type}")

        if resource_id not in self.resources:
            self.resources[resource_id] = {
                "resource_id": resource_id,
                "subject_id": subject_id,
                "concept_id": concept_id,
                "resource_title": title,
                "resource_type": resource_type,
                "difficulty": "medium",
                "url": "",
                "source_sheet": source_sheet
            }

        return resource_id

    def add_activity(
        self,
        student_id,
        subject_id,
        resource_id="",
        concept_id="",
        activity_type="",
        activity_title="",
        status="",
        completion_rate="",
        time_spent_minutes="",
        date="",
        source_sheet=""
    ):
        if not student_id or not activity_type:
            return

        activity_id = stable_id(
            "ACT",
            f"{student_id}::{subject_id}::{resource_id}::{activity_type}::{activity_title}::{date}::{source_sheet}"
        )

        self.activities[activity_id] = {
            "activity_id": activity_id,
            "student_id": student_id,
            "subject_id": subject_id,
            "resource_id": resource_id,
            "concept_id": concept_id,
            "activity_type": activity_type,
            "activity_title": activity_title,
            "status": status,
            "completion_rate": completion_rate,
            "time_spent_minutes": time_spent_minutes,
            "date": date,
            "source_sheet": source_sheet
        }

    def load_static_concepts_and_relations(self):
        for concept in CONCEPT_CATALOG:
            concept_id = concept["concept_id"]
            self.concepts[concept_id] = {
                "concept_id": concept_id,
                "subject_scope": "shared_ai_courses",
                "concept_name": concept["concept_name"],
                "chapter": concept["chapter"],
                "description": concept["description"]
            }

        for relation_id, source, target, relation_type, description in RELATION_CATALOG:
            self.relations[relation_id] = {
                "relation_id": relation_id,
                "source_concept_id": source,
                "target_concept_id": target,
                "relation_type": relation_type,
                "description": description
            }


def parse_simple_student_sheet(sheet, data, subject_id, source_file):
    metadata = get_metadata(sheet)

    for row in range(4, sheet.max_row + 1):
        student_id = detect_student_id(sheet, row, header_row=3)

        if not student_id:
            continue

        group_name = detect_group(sheet, row, header_row=3)
        major = detect_major(sheet, row, header_row=3)

        data.add_student(
            student_id=student_id,
            subject_id=subject_id,
            group_name=group_name,
            major=major,
            source_file=source_file
        )


def parse_progress_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    task_col = find_column(sheet, 3, ["任务完成数"])
    completion_col = find_column(sheet, 3, ["任务点完成百分比"])
    video_col = find_column(sheet, 3, ["视频观看时长"])
    discussion_col = find_column(sheet, 3, ["讨论数"])

    for row in range(4, sheet.max_row + 1):
        student_id = detect_student_id(sheet, row, header_row=3)

        if not student_id:
            continue

        if completion_col:
            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                activity_type="progress_summary",
                activity_title="Task completion summary",
                completion_rate=parse_rate(sheet.cell(row=row, column=completion_col).value),
                status=get_cell(sheet, row, task_col) if task_col else "",
                source_sheet=sheet.title
            )

        if video_col:
            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                activity_type="video_summary",
                activity_title="Video watch summary",
                time_spent_minutes=parse_number(sheet.cell(row=row, column=video_col).value),
                source_sheet=sheet.title
            )

        if discussion_col:
            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                activity_type="discussion_summary",
                activity_title="Discussion participation summary",
                status=get_cell(sheet, row, discussion_col),
                source_sheet=sheet.title
            )


def parse_score_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    headers = {
        get_cell(sheet, 3, col): col
        for col in range(1, sheet.max_column + 1)
        if get_cell(sheet, 3, col)
    }

    for title, col in headers.items():
        if title in ["序号", "学生姓名", "姓名", "学号/工号", "学校", "部门", "院系", "专业", "班级", "行政班级"]:
            continue

        if "综合成绩" not in title and "作业" not in title:
            continue

        work_id = data.add_work(
            subject_id=subject_id,
            title=title,
            work_type="overall",
            source_sheet=sheet.title,
            max_score=100
        )

        concept_id = "C_AI_OVERVIEW"

        for row in range(4, sheet.max_row + 1):
            student_id = detect_student_id(sheet, row, header_row=3)

            if not student_id:
                continue

            score = sheet.cell(row=row, column=col).value

            data.add_answer(
                student_id=student_id,
                subject_id=subject_id,
                work_id=work_id,
                concept_id=concept_id,
                score=score,
                max_score=100,
                status="",
                submitted_at="",
                source_sheet=sheet.title
            )


def parse_wide_work_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    work_type = work_type_for_sheet(sheet.title)
    current_title = ""

    for col in range(1, sheet.max_column + 1):
        row3_value = get_cell(sheet, 3, col)
        row4_value = get_cell(sheet, 4, col)

        if row3_value:
            current_title = row3_value

        if row4_value != "成绩":
            continue

        if not current_title:
            continue

        title = current_title
        concept_id = concept_for_text(title)

        work_id = data.add_work(
            subject_id=subject_id,
            title=title,
            work_type=work_type,
            source_sheet=sheet.title,
            max_score=100
        )

        score_col = col
        submitted_col = col + 1 if col + 1 <= sheet.max_column else None
        status_col = col + 2 if col + 2 <= sheet.max_column else None

        for row in range(5, sheet.max_row + 1):
            student_id = detect_student_id(sheet, row, header_row=3)

            if not student_id:
                continue

            score = sheet.cell(row=row, column=score_col).value
            submitted_at = get_cell(sheet, row, submitted_col) if submitted_col else ""
            status = get_cell(sheet, row, status_col) if status_col else ""

            if score == "" and not status and not submitted_at:
                continue

            data.add_answer(
                student_id=student_id,
                subject_id=subject_id,
                work_id=work_id,
                concept_id=concept_id,
                score=score,
                max_score=100,
                status=status,
                submitted_at=submitted_at,
                source_sheet=sheet.title
            )


def parse_task_completion_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    current_category = ""
    current_resource = ""

    for col in range(7, sheet.max_column + 1):
        category = get_cell(sheet, 3, col)
        resource_title = get_cell(sheet, 4, col)
        metric = get_cell(sheet, 5, col)

        if category:
            current_category = category

        if resource_title:
            current_resource = resource_title

        if metric != "状态":
            continue

        if not current_resource:
            continue

        full_title = current_resource
        resource_id = data.add_resource(subject_id, full_title, sheet.title)
        concept_id = concept_for_text(full_title)

        status_col = col
        date_col = col + 1 if col + 1 <= sheet.max_column else None

        for row in range(6, sheet.max_row + 1):
            student_id = detect_student_id(sheet, row, header_row=3)

            if not student_id:
                continue

            status = get_cell(sheet, row, status_col)
            date = get_cell(sheet, row, date_col) if date_col else ""

            if not status and not date:
                continue

            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                resource_id=resource_id,
                concept_id=concept_id,
                activity_type="task_completion",
                activity_title=full_title,
                status=status,
                completion_rate="1" if is_completed_from_status(status) == "true" else "0",
                date=date,
                source_sheet=sheet.title
            )


def parse_video_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    current_topic = ""
    current_video = ""

    for col in range(7, sheet.max_column + 1):
        topic = get_cell(sheet, 3, col)
        video = get_cell(sheet, 4, col)
        metric = get_cell(sheet, 5, col)

        if topic:
            current_topic = topic

        if video:
            current_video = video

        if metric != "观看时长":
            continue

        if not current_video:
            continue

        resource_title = current_video
        resource_id = data.add_resource(subject_id, resource_title, sheet.title)
        concept_id = concept_for_text(resource_title)

        time_col = col
        start_col = col - 3 if col - 3 >= 1 else None
        end_col = col - 2 if col - 2 >= 1 else None
        rate_col = col - 1 if col - 1 >= 1 else None

        for row in range(6, sheet.max_row + 1):
            student_id = detect_student_id(sheet, row, header_row=3)

            if not student_id:
                continue

            time_spent = parse_number(sheet.cell(row=row, column=time_col).value)
            completion_rate = parse_rate(sheet.cell(row=row, column=rate_col).value) if rate_col else ""
            start_date = get_cell(sheet, row, start_col) if start_col else ""
            end_date = get_cell(sheet, row, end_col) if end_col else ""

            if time_spent == "" and completion_rate == "":
                continue

            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                resource_id=resource_id,
                concept_id=concept_id,
                activity_type="video_watch",
                activity_title=resource_title,
                status="watched" if time_spent and time_spent > 0 else "not_watched",
                completion_rate=completion_rate,
                time_spent_minutes=time_spent,
                date=end_date or start_date,
                source_sheet=sheet.title
            )


def parse_discussion_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    total_col = find_column(sheet, 3, ["总讨论数"])
    post_col = find_column(sheet, 3, ["发表讨论"])
    reply_col = find_column(sheet, 3, ["回复讨论"])
    likes_col = find_column(sheet, 3, ["获赞数"])

    for row in range(4, sheet.max_row + 1):
        student_id = detect_student_id(sheet, row, header_row=3)

        if not student_id:
            continue

        total = get_cell(sheet, row, total_col) if total_col else ""

        data.add_activity(
            student_id=student_id,
            subject_id=subject_id,
            activity_type="discussion",
            activity_title="Discussion participation",
            status=f"total={total}; posts={get_cell(sheet, row, post_col) if post_col else ''}; replies={get_cell(sheet, row, reply_col) if reply_col else ''}; likes={get_cell(sheet, row, likes_col) if likes_col else ''}",
            source_sheet=sheet.title
        )


def parse_attendance_sheet(sheet, data, subject_id, source_file):
    parse_simple_student_sheet(sheet, data, subject_id, source_file)

    for col in range(6, sheet.max_column + 1):
        session_title = get_cell(sheet, 3, col)

        if not session_title:
            continue

        for row in range(5, sheet.max_row + 1):
            student_id = detect_student_id(sheet, row, header_row=3)

            if not student_id:
                continue

            status = get_cell(sheet, row, col)

            if not status:
                continue

            data.add_activity(
                student_id=student_id,
                subject_id=subject_id,
                activity_type="attendance",
                activity_title=session_title,
                status=status,
                completion_rate="1" if is_completed_from_status(status) == "true" else "0",
                date=session_title,
                source_sheet=sheet.title
            )


def parse_workbook(path, data):
    workbook = load_workbook(path, data_only=True, read_only=False)
    source_file = str(path.relative_to(ANON_DIR))

    first_sheet = workbook[workbook.sheetnames[0]]
    metadata = get_metadata(first_sheet)
    subject_id = data.add_subject(source_file, metadata)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        if sheet_name in ["综合成绩", "综合成绩（各权重项百分制得分）"]:
            parse_score_sheet(sheet, data, subject_id, source_file)

        elif sheet_name == "学生学习进度详情":
            parse_progress_sheet(sheet, data, subject_id, source_file)

        elif sheet_name in ["作业统计", "考试统计", "章节测验统计", "线下成绩统计"]:
            parse_wide_work_sheet(sheet, data, subject_id, source_file)

        elif sheet_name == "任务点完成详情":
            parse_task_completion_sheet(sheet, data, subject_id, source_file)

        elif sheet_name == "音视频观看详情":
            parse_video_sheet(sheet, data, subject_id, source_file)

        elif sheet_name == "讨论参与详情":
            parse_discussion_sheet(sheet, data, subject_id, source_file)

        elif sheet_name == "签到详情统计":
            parse_attendance_sheet(sheet, data, subject_id, source_file)

    return source_file


def write_outputs(data):
    write_csv(
        OUTPUT_FILES["students"],
        sorted(data.students.values(), key=lambda row: (row["subject_id"], row["student_id"])),
        ["student_id", "subject_id", "group_name", "major", "source_file"]
    )

    write_csv(
        OUTPUT_FILES["subjects"],
        sorted(data.subjects.values(), key=lambda row: row["subject_id"]),
        ["subject_id", "subject_name", "teacher", "class_name", "source_file", "description"]
    )

    write_csv(
        OUTPUT_FILES["works"],
        sorted(data.works.values(), key=lambda row: (row["subject_id"], row["work_type"], row["work_title"])),
        ["work_id", "subject_id", "work_title", "work_type", "date", "max_score", "source_sheet"]
    )

    write_csv(
        OUTPUT_FILES["answers"],
        sorted(data.answers.values(), key=lambda row: (row["student_id"], row["work_id"])),
        [
            "answer_id", "student_id", "subject_id", "work_id", "concept_id",
            "score", "max_score", "score_rate", "is_correct", "status",
            "is_completed", "submitted_at", "source_sheet"
        ]
    )

    write_csv(
        OUTPUT_FILES["activities"],
        sorted(data.activities.values(), key=lambda row: (row["student_id"], row["activity_type"], row["activity_title"])),
        [
            "activity_id", "student_id", "subject_id", "resource_id", "concept_id",
            "activity_type", "activity_title", "status", "completion_rate",
            "time_spent_minutes", "date", "source_sheet"
        ]
    )

    write_csv(
        OUTPUT_FILES["concepts"],
        sorted(data.concepts.values(), key=lambda row: row["concept_id"]),
        ["concept_id", "subject_scope", "concept_name", "chapter", "description"]
    )

    write_csv(
        OUTPUT_FILES["relations"],
        sorted(data.relations.values(), key=lambda row: row["relation_id"]),
        ["relation_id", "source_concept_id", "target_concept_id", "relation_type", "description"]
    )

    write_csv(
        OUTPUT_FILES["resources"],
        sorted(data.resources.values(), key=lambda row: row["resource_id"]),
        [
            "resource_id", "subject_id", "concept_id", "resource_title",
            "resource_type", "difficulty", "url", "source_sheet"
        ]
    )


def write_report(data, processed_files):
    lines = []

    lines.append("# Real Course Data Normalization Report")
    lines.append("")
    lines.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## Purpose")
    lines.append("")
    lines.append(
        "This report summarizes the transformation of anonymized course exports "
        "into simple normalized tables used by the teaching diagnosis system."
    )
    lines.append("")
    lines.append("## Normalized vocabulary")
    lines.append("")
    lines.append("- `student`: anonymized learner")
    lines.append("- `subject`: course or class")
    lines.append("- `work`: quiz, homework, exam, chapter test or global score")
    lines.append("- `answer`: student result on a work")
    lines.append("- `activity`: learning behavior such as video watching, task completion, discussion or attendance")
    lines.append("- `concept`: course knowledge point")
    lines.append("- `relation`: link between concepts")
    lines.append("- `resource`: lesson, video, PDF, task or exercise")
    lines.append("- `diagnosis`: generated later from answers, activities, concepts and relations")
    lines.append("")
    lines.append("## Files processed")
    lines.append("")

    for file in processed_files:
        lines.append(f"- `{file}`")

    lines.append("")
    lines.append("## Output tables")
    lines.append("")
    lines.append(f"- Students: **{len(data.students)}**")
    lines.append(f"- Subjects: **{len(data.subjects)}**")
    lines.append(f"- Works: **{len(data.works)}**")
    lines.append(f"- Answers: **{len(data.answers)}**")
    lines.append(f"- Activities: **{len(data.activities)}**")
    lines.append(f"- Concepts: **{len(data.concepts)}**")
    lines.append(f"- Relations: **{len(data.relations)}**")
    lines.append(f"- Resources: **{len(data.resources)}**")
    lines.append("")
    lines.append("## Privacy note")
    lines.append("")
    lines.append(
        "The normalized files are generated from anonymized files. They still describe student learning behavior, "
        "so they remain local and are ignored by Git."
    )

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")


def main():
    if not ANON_DIR.exists():
        raise FileNotFoundError(
            f"Anonymized directory not found: {ANON_DIR}. Run scripts/anonymize_real_course_data.py first."
        )

    xlsx_files = sorted(ANON_DIR.rglob("*.xlsx"))

    if not xlsx_files:
        raise FileNotFoundError(
            f"No anonymized Excel files found in {ANON_DIR}. Run scripts/anonymize_real_course_data.py first."
        )

    data = NormalizedData()
    data.load_static_concepts_and_relations()

    processed_files = []

    for path in xlsx_files:
        print(f"Normalizing {path.relative_to(ANON_DIR)}...")
        processed_file = parse_workbook(path, data)
        processed_files.append(processed_file)

    write_outputs(data)
    write_report(data, processed_files)

    print("")
    print("Normalization completed.")
    print(f"Students: {len(data.students)}")
    print(f"Subjects: {len(data.subjects)}")
    print(f"Works: {len(data.works)}")
    print(f"Answers: {len(data.answers)}")
    print(f"Activities: {len(data.activities)}")
    print(f"Concepts: {len(data.concepts)}")
    print(f"Relations: {len(data.relations)}")
    print(f"Resources: {len(data.resources)}")
    print("")
    print(f"Normalized files written to: {NORMALIZED_DIR}")
    print(f"Report written to: {REPORT_FILE}")


if __name__ == "__main__":
    main()