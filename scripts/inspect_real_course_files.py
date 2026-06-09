from pathlib import Path
from io import BytesIO
from datetime import datetime
import zipfile
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "real_courses"
INVENTORY_DIR = ROOT / "data" / "processed" / "real_courses" / "inventory"
INVENTORY_DIR.mkdir(parents=True, exist_ok=True)

MARKDOWN_OUTPUT = INVENTORY_DIR / "real_course_inventory.md"
JSON_OUTPUT = INVENTORY_DIR / "real_course_inventory.json"


SENSITIVE_KEYWORDS = [
    "姓名", "名字", "学生姓名", "学号", "学生号", "工号", "账号", "用户",
    "手机号", "手机", "电话", "邮箱", "身份证", "证件",
    "name", "student name", "student_id", "student id", "id", "email", "phone",
    "number", "username", "user"
]


def is_sensitive_column(column_name):
    if column_name is None:
        return False

    text = str(column_name).strip().lower()

    return any(keyword.lower() in text for keyword in SENSITIVE_KEYWORDS)


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip()


def inspect_workbook(workbook, source_name):
    result = {
        "source": source_name,
        "type": "excel",
        "sheets": []
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        headers = []
        if sheet.max_row >= 1:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [normalize_header(value) for value in first_row]

        sensitive_columns = [
            column for column in headers
            if is_sensitive_column(column)
        ]

        result["sheets"].append({
            "sheet_name": sheet_name,
            "rows": max(sheet.max_row - 1, 0),
            "columns_count": sheet.max_column,
            "columns": headers,
            "sensitive_columns_detected": sensitive_columns,
            "contains_possible_personal_data": len(sensitive_columns) > 0
        })

    return result


def inspect_xlsx_file(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    return inspect_workbook(workbook, path.name)


def inspect_xlsx_bytes(content, source_name):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    return inspect_workbook(workbook, source_name)


def inspect_zip_file(path):
    result = {
        "source": path.name,
        "type": "zip",
        "entries": []
    }

    with zipfile.ZipFile(path, "r") as archive:
        for entry in archive.namelist():
            if entry.endswith("/"):
                continue

            entry_info = {
                "name": entry,
                "extension": Path(entry).suffix.lower()
            }

            if entry.lower().endswith(".xlsx"):
                try:
                    content = archive.read(entry)
                    entry_info["excel_inventory"] = inspect_xlsx_bytes(
                        content,
                        f"{path.name}::{entry}"
                    )
                except Exception as error:
                    entry_info["error"] = str(error)

            result["entries"].append(entry_info)

    return result


def inspect_file(path):
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return inspect_xlsx_file(path)

    if suffix == ".zip":
        return inspect_zip_file(path)

    return {
        "source": path.name,
        "type": suffix.replace(".", "") or "unknown",
        "note": "File type registered but not deeply inspected by this script."
    }


def write_markdown(inventory):
    lines = []

    lines.append("# Real Course Data Inventory")
    lines.append("")
    lines.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## Purpose")
    lines.append("")
    lines.append(
        "This inventory lists the raw course files, Excel sheets, columns and possible sensitive fields. "
        "It does not include actual student values."
    )
    lines.append("")
    lines.append("## Safety rule")
    lines.append("")
    lines.append(
        "Raw files must stay local in `data/raw/real_courses/` and must never be committed to Git."
    )
    lines.append("")

    if not inventory:
        lines.append("No files were found in `data/raw/real_courses/`.")
        lines.append("")
    else:
        lines.append("## Files inspected")
        lines.append("")

    for item in inventory:
        lines.append(f"### {item['source']}")
        lines.append("")
        lines.append(f"- Type: `{item['type']}`")
        lines.append("")

        if item["type"] == "excel":
            write_excel_section(lines, item)

        elif item["type"] == "zip":
            lines.append("#### ZIP entries")
            lines.append("")

            for entry in item["entries"]:
                lines.append(f"- `{entry['name']}` ({entry['extension']})")

                if "error" in entry:
                    lines.append(f"  - Error while inspecting: `{entry['error']}`")

                if "excel_inventory" in entry:
                    excel_inventory = entry["excel_inventory"]
                    lines.append("")
                    lines.append(f"#### Excel inside ZIP: `{entry['name']}`")
                    lines.append("")
                    write_excel_section(lines, excel_inventory)

            lines.append("")

        else:
            lines.append(f"- Note: {item.get('note', '')}")
            lines.append("")

    MARKDOWN_OUTPUT.write_text("\n".join(lines), encoding="utf-8")


def write_excel_section(lines, excel_inventory):
    sheets = excel_inventory.get("sheets", [])

    lines.append(f"- Sheets detected: **{len(sheets)}**")
    lines.append("")

    for sheet in sheets:
        lines.append(f"#### Sheet: `{sheet['sheet_name']}`")
        lines.append("")
        lines.append(f"- Rows: **{sheet['rows']}**")
        lines.append(f"- Columns: **{sheet['columns_count']}**")
        lines.append(f"- Possible personal data: **{sheet['contains_possible_personal_data']}**")

        sensitive_columns = sheet["sensitive_columns_detected"]

        if sensitive_columns:
            lines.append(f"- Sensitive columns detected: `{', '.join(sensitive_columns)}`")
        else:
            lines.append("- Sensitive columns detected: none")

        lines.append("")
        lines.append("Columns:")
        lines.append("")

        for column in sheet["columns"]:
            marker = " ⚠️" if column in sensitive_columns else ""
            lines.append(f"- `{column}`{marker}")

        lines.append("")


def main():
    if not RAW_DIR.exists():
        raise FileNotFoundError(f"Raw directory not found: {RAW_DIR}")

    files = [
        path for path in RAW_DIR.iterdir()
        if path.is_file()
    ]

    inventory = []

    for path in files:
        print(f"Inspecting {path.name}...")
        try:
            inventory.append(inspect_file(path))
        except Exception as error:
            inventory.append({
                "source": path.name,
                "type": path.suffix.lower().replace(".", ""),
                "error": str(error)
            })

    JSON_OUTPUT.write_text(
        json.dumps(inventory, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    write_markdown(inventory)

    print("")
    print(f"Inventory written to: {MARKDOWN_OUTPUT}")
    print(f"JSON inventory written to: {JSON_OUTPUT}")


if __name__ == "__main__":
    main()